"""
reader.py — 讀取來源 Excel（訂單資訊管制表）。

使用 openpyxl 的 read_only + data_only 模式開啟，
確保絕對不會修改原始檔案。
"""

import logging

import openpyxl
import pandas as pd


# 來源 Excel 中必須存在的欄位名稱
REQUIRED_COLUMNS = {"Customer", "品名 \n規格", "品號", "客戶交期", "未交量"}


def read_source(excel_path: str, sheet_name) -> pd.DataFrame:
    """
    以唯讀模式讀取來源 Excel 檔案，回傳原始欄位名稱的 DataFrame。

    Parameters
    ----------
    excel_path : str
        訂單資訊管制表的完整路徑（NAS 或本機）。
    sheet_name : str or int
        工作表名稱或 0-based 索引（預設為「待出貨-膜類」）。

    Raises
    ------
    FileNotFoundError : 檔案不存在或無法存取
    PermissionError   : 檔案被其他程式占用
    ValueError        : 工作表不存在或缺少必要欄位
    """
    logging.info("讀取來源 Excel：%s  工作表：%s", excel_path, sheet_name)

    # read_only=True：不會寫回任何變更
    # data_only=True：回傳公式的計算結果，而非公式本身
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # 依名稱或索引取得工作表
    if isinstance(sheet_name, int):
        ws = wb.worksheets[sheet_name]
    else:
        if sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            wb.close()
            raise ValueError(
                f"找不到工作表「{sheet_name}」。\n可用工作表：{available}"
            )
        ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("工作表是空的，請確認來源檔案內容。")

    # 第一列為標題，其餘為資料
    headers = [str(h) if h is not None else f"_col{i}" for i, h in enumerate(rows[0])]
    data = rows[1:]

    df = pd.DataFrame(data, columns=headers)
    logging.info("讀取完成，共 %d 列（不含標題）", len(df))

    # 檢查必要欄位是否齊全
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise ValueError(
            f"來源 Excel 缺少以下欄位，請確認欄位名稱或聯繫 IT：\n{', '.join(sorted(missing))}"
        )

    return df
