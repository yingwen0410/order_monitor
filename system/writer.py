"""
writer.py — 產出三頁式 Excel 報表（使用 openpyxl）。

工作表：
  1. 過期    — 客戶交期已到或超過今日的訂單，依 (客戶, 品號) 合計
  2. 未過期  — 客戶交期尚未到期的訂單，同上結構
  3. 客戶總表 — 依客戶分組，每個品號展開「過期 / 未過期」兩列
"""

import logging
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from utils import (
    COL_CUSTOMER, COL_PRODUCT_NAME, COL_PART_NO,
    COL_DELIVERY, COL_UNDELIVERED, COL_STOCK, COL_STATUS, COL_SURPLUS,
)


# ---------------------------------------------------------------------------
# 樣式定義
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    """建立純色填充。"""
    return PatternFill("solid", fgColor=hex_color)


FILL_COL_HEADER  = _fill("2C3E50")   # 深色 — 欄位標題列
FILL_CUST_HEADER = _fill("DDDDDD")   # 淺灰 — 客戶段落標題
FILL_WHITE       = _fill("FFFFFF")   # 白色 — 預設背景
FILL_GROUP_ALT   = _fill("F3F3F3")   # 極淺灰 — 交替客戶群組底色
FILL_SURPLUS_POS = _fill("C6EFCE")   # 淺綠 — 庫存充足（差額 > 0）
FILL_SURPLUS_NEG = _fill("FFCCCC")   # 淺紅 — 庫存不足（差額 <= 0）

FONT_COL_HEADER  = Font(name="微軟正黑體", bold=True, color="FFFFFF", size=11)
FONT_CUST_HEADER = Font(name="標楷體", bold=True, color="000000", size=12)
FONT_NORMAL      = Font(name="標楷體", size=12)

THIN_SIDE   = Side(border_style="thin",   color="D0D0D0")
MEDIUM_SIDE = Side(border_style="medium", color="AAAAAA")
BORDER      = Border(left=THIN_SIDE, right=THIN_SIDE,
                     top=THIN_SIDE, bottom=THIN_SIDE)

LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")

NUM_FORMAT = "#,##0"   # 千分位、無小數


# ---------------------------------------------------------------------------
# 輔助函式
# ---------------------------------------------------------------------------

def _set_cell(ws, row: int, col: int, value, font=None, fill=None,
              alignment=None, border=None, number_format=None):
    """設定單一儲存格的值與樣式。"""
    cell = ws.cell(row=row, column=col, value=value)
    if font:          cell.font          = font
    if fill:          cell.fill          = fill
    if alignment:     cell.alignment     = alignment
    if border:        cell.border        = border
    if number_format: cell.number_format = number_format
    return cell


def _write_num(ws, row: int, col: int, value, fill, border):
    """寫入數值儲存格（靠右對齊 + 千分位格式），或顯示 'N/A'。"""
    is_num = isinstance(value, (int, float)) and not (
        isinstance(value, float) and pd.isna(value)
    )
    _set_cell(ws, row, col, value,
              font=FONT_NORMAL,
              fill=fill,
              alignment=RIGHT if is_num else LEFT,
              border=border,
              number_format=NUM_FORMAT if is_num else None)


def _auto_width(ws, min_width=10, max_width=42):
    """自動調整所有欄位寬度。"""
    for col_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = (
            max(min_width, min(length + 2, max_width))
        )


def _write_sheet_header(ws, headers: list[str]):
    """寫入工作表標題列（第 1 列），並凍結窗格。"""
    for col_idx, title in enumerate(headers, start=1):
        _set_cell(ws, 1, col_idx, title,
                  font=FONT_COL_HEADER, fill=FILL_COL_HEADER,
                  alignment=CENTER, border=BORDER)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _fmt_date(ts) -> str:
    """將 Timestamp 格式化為 YYYY/MM/DD 字串。"""
    return ts.strftime("%Y/%m/%d") if pd.notna(ts) else ""


def _surplus_fill(value) -> PatternFill:
    """根據庫存差額決定儲存格底色（綠=充足，紅=不足）。"""
    if not isinstance(value, (int, float)) or pd.isna(value):
        return FILL_WHITE
    return FILL_SURPLUS_POS if value > 0 else FILL_SURPLUS_NEG


def _stock_val(v):
    """庫存量顯示值：NaN 顯示為 'N/A'。"""
    return "N/A" if (not isinstance(v, (int, float))) or pd.isna(v) else v


def _surplus_val(v):
    """庫存差額顯示值：NaN 顯示為 'N/A'。"""
    return "N/A" if (not isinstance(v, (int, float))) or pd.isna(v) else v


# ---------------------------------------------------------------------------
# 共用常數
# ---------------------------------------------------------------------------

# R&D 品號：庫存量顯示「人員自行查找」、庫存差額留空
_RD_PART_NOS = {"TC-ROL", "TC-PC"}


# ---------------------------------------------------------------------------
# 累積庫存計算（跨客戶、依交期優先序）
# ---------------------------------------------------------------------------

def _calc_surplus_map(df: pd.DataFrame, today: date) -> dict:
    """供「客戶總表」使用：計算每個 (客戶, 狀態, 品號) 合計後的累積庫存差額。"""
    df = df.copy()
    df["_status"] = df[COL_DELIVERY].apply(
        lambda d: "過期" if d.date() <= today else "未過期"
    )

    agg = df.groupby([COL_CUSTOMER, "_status", COL_PART_NO], as_index=False).agg(**{
        COL_UNDELIVERED: (COL_UNDELIVERED, "sum"),
        COL_STOCK:       (COL_STOCK,       "first"),
        COL_DELIVERY:    (COL_DELIVERY,    "min"),
    })

    agg["_so"] = agg["_status"].map({"過期": 0, "未過期": 1})
    agg = agg.sort_values([COL_PART_NO, "_so", COL_DELIVERY]).reset_index(drop=True)
    agg["_cum"] = agg.groupby(COL_PART_NO)[COL_UNDELIVERED].cumsum()

    agg["_avail"]   = agg[COL_STOCK] - (agg["_cum"] - agg[COL_UNDELIVERED])
    agg["_surplus"] = agg[COL_STOCK] - agg["_cum"]

    return {
        (str(row[COL_CUSTOMER]), str(row["_status"]), str(row[COL_PART_NO])): {
            "avail":   row["_avail"],
            "surplus": row["_surplus"],
        }
        for _, row in agg.iterrows()
    }


def _calc_global_stock(df: pd.DataFrame, today: date) -> pd.DataFrame:
    """供「明細 sheet」使用：計算每一筆訂單（不合併）依時序扣除的可用庫存與差額。"""
    df = df.copy()
    df["_status"] = df[COL_DELIVERY].apply(
        lambda d: "過期" if d.date() <= today else "未過期"
    )
    df["_so"] = df["_status"].map({"過期": 0, "未過期": 1})
    
    # 排序：品號 -> 狀態(過期優先) -> 交期
    df.sort_values([COL_PART_NO, "_so", COL_DELIVERY], inplace=True)
    
    df["_cum"] = df.groupby(COL_PART_NO)[COL_UNDELIVERED].cumsum()
    df["_avail"]   = df[COL_STOCK] - (df["_cum"] - df[COL_UNDELIVERED])
    df["_surplus"] = df[COL_STOCK] - df["_cum"]
    
    return df


def _calc_final_balance(df: pd.DataFrame, today: date) -> pd.DataFrame:
    """
    計算每個品號在全部客戶訂單指派後的「最終剩餘庫存」。
    未交量拆為過期與未過期兩欄。
    """
    df = df.copy()
    df["_expired"] = df[COL_DELIVERY].dt.date <= today

    # 過期未交量 / 未過期未交量
    exp_agg = df[df["_expired"]].groupby(COL_PART_NO)[COL_UNDELIVERED].sum().rename("exp_qty")
    ne_agg  = df[~df["_expired"]].groupby(COL_PART_NO)[COL_UNDELIVERED].sum().rename("ne_qty")

    agg = df.groupby(COL_PART_NO, as_index=False).agg(**{
        COL_PRODUCT_NAME: (COL_PRODUCT_NAME, "first"),
        COL_UNDELIVERED:  (COL_UNDELIVERED,  "sum"),
        COL_STOCK:        (COL_STOCK,        "first"),
    })
    agg = agg.join(exp_agg, on=COL_PART_NO).join(ne_agg, on=COL_PART_NO)
    agg["exp_qty"] = agg["exp_qty"].fillna(0)
    agg["ne_qty"]  = agg["ne_qty"].fillna(0)
    agg["final_balance"] = agg[COL_STOCK] - agg[COL_UNDELIVERED]
    agg.sort_values(COL_PART_NO, inplace=True)
    agg.reset_index(drop=True, inplace=True)
    return agg


# ---------------------------------------------------------------------------
# Sheet 1 & 2：過期 / 未過期
# ---------------------------------------------------------------------------

_STATUS_COLUMNS = ["客戶交期", "客戶", "品名規格", "品號", "未交量", "庫存量", "庫存差額"]


def _build_status_df(df: pd.DataFrame, expired: bool, today: date) -> pd.DataFrame:
    """依過期 / 未過期篩選，列出個別明細，並依交期排序（由舊至新）。"""
    mask = df[COL_DELIVERY].dt.date <= today if expired else df[COL_DELIVERY].dt.date > today
    sub  = df[mask].copy()
    if sub.empty:
        return pd.DataFrame(columns=_STATUS_COLUMNS)

    sub.sort_values(COL_DELIVERY, inplace=True)
    sub.reset_index(drop=True, inplace=True)
    return sub


def _write_status_sheet(ws, agg_df: pd.DataFrame, surplus_map: dict, status_name: str):
    """將合計後的資料寫入過期或未過期工作表。"""
    _write_sheet_header(ws, _STATUS_COLUMNS)

    # 依客戶分組交替底色
    unique_customers = list(dict.fromkeys(agg_df[COL_CUSTOMER]))
    customer_fill = {
        c: (FILL_WHITE if i % 2 == 0 else FILL_GROUP_ALT)
        for i, c in enumerate(unique_customers)
    }

    prev_customer = None
    for i, row in agg_df.iterrows():
        r        = i + 2
        customer = row[COL_CUSTOMER]
        gfill    = customer_fill[customer]
        # 客戶切換時加粗上邊框
        top      = MEDIUM_SIDE if customer != prev_customer else THIN_SIDE
        bdr      = Border(left=THIN_SIDE, right=THIN_SIDE, top=top, bottom=THIN_SIDE)
        stock    = row[COL_STOCK]

        # 文字欄位
        for col_idx, val in enumerate(
            [_fmt_date(row[COL_DELIVERY]), customer,
             row[COL_PRODUCT_NAME], row[COL_PART_NO]],
            start=1
        ):
            _set_cell(ws, r, col_idx, val, font=FONT_NORMAL, fill=gfill,
                      alignment=LEFT, border=bdr)

        # 取得該筆訂單計算好的行級別可用庫存與差額
        is_rd   = row[COL_PART_NO] in _RD_PART_NOS
        avail   = row.get("_avail", stock)
        surplus = row.get("_surplus", stock - row[COL_UNDELIVERED])

        # 數值欄位（5=未交量, 6=庫存量, 7=庫存差額）
        _write_num(ws, r, 5, row[COL_UNDELIVERED], gfill, bdr)
        if is_rd:
            # R&D 品號：庫存量顯示「人員自行查找」，差額留空
            _set_cell(ws, r, 6, "人員自行查找", font=FONT_NORMAL,
                      fill=gfill, alignment=LEFT, border=bdr)
            _set_cell(ws, r, 7, None, fill=gfill, border=bdr)
        else:
            _write_num(ws, r, 6, _stock_val(avail),   gfill, bdr)
            _write_num(ws, r, 7, _surplus_val(surplus), _surplus_fill(surplus), bdr)

        prev_customer = customer

    _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet 3：客戶總表
# ---------------------------------------------------------------------------

_SUMMARY_COLS = ["客戶", "品號", "最大安基量", "品名規格",
                 "狀態", "未交量合計", "庫存量", "庫存差額"]
_STATUS_ORDER = {"過期": 0, "未過期": 1}


def _build_summary_df(df: pd.DataFrame, today: date) -> pd.DataFrame:
    """建立客戶總表的 DataFrame：依 (客戶, 狀態, 品號) 合計（差額由 surplus_map 提供）。"""
    df = df.copy()
    df[COL_STATUS] = df[COL_DELIVERY].apply(
        lambda d: "過期" if d.date() <= today else "未過期"
    )
    agg = df.groupby([COL_CUSTOMER, COL_STATUS, COL_PART_NO], as_index=False).agg(**{
        COL_PRODUCT_NAME: (COL_PRODUCT_NAME, "first"),
        COL_UNDELIVERED:  (COL_UNDELIVERED,  "sum"),
        COL_STOCK:        (COL_STOCK,         "first"),
    })
    agg["_so"] = agg[COL_STATUS].map(_STATUS_ORDER)
    agg.sort_values([COL_CUSTOMER, COL_PART_NO, "_so"], inplace=True)
    agg.drop(columns=["_so"], inplace=True)
    agg.reset_index(drop=True, inplace=True)
    return agg


def _write_summary_sheet(ws, summary_df: pd.DataFrame, allow_lookup: dict, surplus_map: dict):
    """將客戶總表寫入工作表，包含最大安基量與合併儲存格。"""
    _write_sheet_header(ws, _SUMMARY_COLS)
    # 依客戶分組交替底色
    unique_customers = list(dict.fromkeys(summary_df[COL_CUSTOMER]))
    customer_fill = {
        c: (FILL_WHITE if i % 2 == 0 else FILL_GROUP_ALT)
        for i, c in enumerate(unique_customers)
    }

    r = 2

    for customer, cust_df in summary_df.groupby(COL_CUSTOMER, sort=False):
        gfill = customer_fill[customer]

        # ── 客戶標題列（A 欄顯示客戶名稱，其餘欄灰底留空）──
        _set_cell(ws, r, 1, customer,
                  font=FONT_CUST_HEADER, fill=FILL_CUST_HEADER,
                  alignment=LEFT, border=BORDER)
        for c in range(2, len(_SUMMARY_COLS) + 1):
            _set_cell(ws, r, c, None, fill=FILL_CUST_HEADER, border=BORDER)
        ws.row_dimensions[r].height = 20
        r += 1

        # ── 依品號展開明細列 ──
        for part_no, part_df in cust_df.groupby(COL_PART_NO, sort=False):
            rows      = list(part_df.itertuples(index=False))
            n         = len(rows)
            start_r   = r
            prod_name = rows[0].__getattribute__(COL_PRODUCT_NAME)
            is_rd     = part_no in _RD_PART_NOS
            max_allow = allow_lookup.get((customer, part_no))

            for row_data in rows:
                stock  = row_data.__getattribute__(COL_STOCK)
                status = row_data.__getattribute__(COL_STATUS)
                undelivered = row_data.__getattribute__(COL_UNDELIVERED)
                # 從累積庫存表取得（區分客戶、狀態、品號）
                cum     = surplus_map.get((str(customer), str(status), str(part_no)), {})
                avail   = cum.get("avail",   stock)
                surplus = cum.get("surplus", stock - undelivered)

                # A 欄：留空（屬於客戶群組）
                _set_cell(ws, r, 1, None, fill=gfill, border=BORDER)

                # B=品號, C=最大安基量, D=品名規格（僅首列填值，多列時合併）
                if r == start_r:
                    _set_cell(ws, r, 2, part_no, font=FONT_NORMAL, fill=gfill,
                              alignment=LEFT, border=BORDER)
                    _set_cell(ws, r, 3, max_allow, font=FONT_NORMAL, fill=gfill,
                              alignment=LEFT, border=BORDER)
                    _set_cell(ws, r, 4, prod_name, font=FONT_NORMAL, fill=gfill,
                              alignment=LEFT, border=BORDER)
                else:
                    _set_cell(ws, r, 2, None, fill=gfill, border=BORDER)
                    _set_cell(ws, r, 3, None, fill=gfill, border=BORDER)
                    _set_cell(ws, r, 4, None, fill=gfill, border=BORDER)

                # E=狀態
                _set_cell(ws, r, 5, status, font=FONT_NORMAL, fill=gfill,
                          alignment=LEFT, border=BORDER)
                # F=未交量合計
                _write_num(ws, r, 6, row_data.__getattribute__(COL_UNDELIVERED),
                           gfill, BORDER)
                # G=庫存量（R&D 品號顯示「人員自行查找」）
                if is_rd:
                    _set_cell(ws, r, 7, "人員自行查找", font=FONT_NORMAL,
                              fill=gfill, alignment=LEFT, border=BORDER)
                else:
                    _write_num(ws, r, 7, _stock_val(avail), gfill, BORDER)
                # H=庫存差額（R&D 品號留空）
                if is_rd:
                    _set_cell(ws, r, 8, None, fill=gfill, border=BORDER)
                else:
                    _write_num(ws, r, 8, _surplus_val(surplus),
                               _surplus_fill(surplus), BORDER)

                r += 1

            # 同一品號有多筆狀態時，合併 B/C/D 欄
            if n > 1:
                for col in (2, 3, 4):
                    ws.merge_cells(start_row=start_r, start_column=col,
                                   end_row=r - 1, end_column=col)
                    ws.cell(start_r, col).alignment = Alignment(
                        horizontal="left", vertical="center"
                    )

    _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


_BALANCE_COLS = ["品號", "品名規格", "原始庫存",
                 "過期未交量", "未過期未交量", "庫存差額"]


def _write_balance_sheet(ws, balance_df: pd.DataFrame):
    """
    寫出【品號庫存】sheet：顯示每個品號將全部客戶訂單指派完後的最終剩餘庫存。
    未交量拆為過期與未過期兩欄，方便一目瞭然尌主要紧張來源。
    """
    _write_sheet_header(ws, _BALANCE_COLS)
    for i, row in balance_df.iterrows():
        r        = i + 2
        part_no  = row[COL_PART_NO]
        is_rd    = part_no in _RD_PART_NOS
        final    = row["final_balance"]
        bdr      = BORDER
        fill     = FILL_WHITE if i % 2 == 0 else FILL_GROUP_ALT

        _set_cell(ws, r, 1, part_no,               font=FONT_NORMAL, fill=fill, alignment=LEFT, border=bdr)
        _set_cell(ws, r, 2, row[COL_PRODUCT_NAME], font=FONT_NORMAL, fill=fill, alignment=LEFT, border=bdr)

        if is_rd:
            # R&D 品號：庫存量人員自行查找，差額留空
            _set_cell(ws, r, 3, "人員自行查找", font=FONT_NORMAL, fill=fill, alignment=LEFT, border=bdr)
            _write_num(ws, r, 4, row["exp_qty"] or None, fill, bdr)
            _write_num(ws, r, 5, row["ne_qty"]  or None, fill, bdr)
            _set_cell(ws, r, 6, None, fill=fill, border=bdr)
        else:
            _write_num(ws, r, 3, _stock_val(row[COL_STOCK]),  fill, bdr)
            _write_num(ws, r, 4, row["exp_qty"] or None,      fill, bdr)
            _write_num(ws, r, 5, row["ne_qty"]  or None,      fill, bdr)
            _write_num(ws, r, 6, _surplus_val(final), _surplus_fill(final), bdr)

    _auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# \u4e3b\u8981\u8f38\u51fa\u51fd\u5f0f
# ---------------------------------------------------------------------------

def write_report(df: pd.DataFrame, today: date, output_path: str,
                 allow_lookup: dict = None):
    """
    產出完整的三頁式報表並儲存。

    Parameters
    ----------
    df : pd.DataFrame     正規化後的訂單資料
    today : date          今日日期（用於判斷過期與否）
    output_path : str     報表儲存路徑
    allow_lookup : dict   (客戶, 品號) → 最大安基量 的對照表
    """
    logging.info("產生報表 → %s", output_path)
    wb = Workbook()
    wb.remove(wb.active)  # 移除預設的空白工作表

    # 預先計算累積庫存差額（供「明細sheet」使用的行級別計算）
    df_global = _calc_global_stock(df, today)

    # 預先計算累積庫存差額（供「客戶總表」使用，確保跨客戶、跨狀態庫存扣除一致）
    surplus_map = _calc_surplus_map(df, today)

    # Sheet 1：過期訂單
    ws_exp = wb.create_sheet("過期")
    expired_df = _build_status_df(df_global, expired=True, today=today)
    _write_status_sheet(ws_exp, expired_df, surplus_map, "過期")
    logging.info("過期 sheet：%d 筆", len(expired_df))

    # Sheet 2：未過期訂單
    ws_ne = wb.create_sheet("未過期")
    notexp_df = _build_status_df(df_global, expired=False, today=today)
    _write_status_sheet(ws_ne, notexp_df, surplus_map, "未過期")
    logging.info("未過期 sheet：%d 筆", len(notexp_df))

    # Sheet 3：客戶總表
    ws_sum = wb.create_sheet("客戶總表")
    summary_df = _build_summary_df(df, today)
    _write_summary_sheet(ws_sum, summary_df, allow_lookup or {}, surplus_map)
    logging.info("客戶總表 sheet：%d 筆（不含標題列）", len(summary_df))

    # Sheet 4：品號庫存總覽（每個品號最終剩餘庫存）
    ws_bal = wb.create_sheet("品號庫存")
    balance_df = _calc_final_balance(df, today)
    _write_balance_sheet(ws_bal, balance_df)
    logging.info("品號庫存 sheet：%d 筆", len(balance_df))

    wb.save(output_path)
    logging.info("報表已儲存")
